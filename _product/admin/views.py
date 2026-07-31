import csv
import logging
from datetime import date
from io import BytesIO

from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from wagtail.admin.auth import user_passes_test
from wagtail.snippets.views.snippets import CreateView

from _product.models import Product, SavingTransaction
from _product.services.saving_workflow import (
    create_saving_transaction,
    record_manual_saving_transaction,
)

logger = logging.getLogger(__name__)


class SavingTransactionCreateView(CreateView):
    def save_instance(self):
        instance = super().save_instance()
        try:
            instance.created_by = self.request.user
            instance.save(update_fields=("created_by", "updated_at"))
            if instance.payment_channel == SavingTransaction.PaymentChannel.MANUAL:
                instance = record_manual_saving_transaction(
                    instance,
                    recorded_by=self.request.user,
                )
                messages.success(
                    self.request,
                    f"Setoran manual {instance.transaction_code} tersimpan sebagai lunas. Email konfirmasi dijadwalkan.",
                )
            else:
                instance = create_saving_transaction(saving_txn=instance)
                messages.success(
                    self.request,
                    f"Payment link {instance.transaction_code} berhasil dibuat: {instance.payment_link_url}",
                )
        except Exception as exc:
            messages.error(self.request, f"Transaksi setoran gagal diselesaikan: {exc}")
            logger.exception("Unexpected error in SavingTransaction create for pk=%s", instance.pk)
        return instance


def _report_permission(user):
    return user.has_perm("wagtailadmin.access_admin")


def _filtered_savings(request):
    today = timezone.localdate()
    default_from = today.replace(day=1)
    date_from = parse_date(request.GET.get("date_from", "")) or default_from
    date_to = parse_date(request.GET.get("date_to", "")) or today
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    status = request.GET.get("status", SavingTransaction.Status.PAID)
    product_id = request.GET.get("product", "")
    qs = SavingTransaction.objects.select_related("product", "created_by").filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    )
    if status in dict(SavingTransaction.Status.choices):
        qs = qs.filter(status=status)
    else:
        status = ""
    if product_id.isdigit():
        qs = qs.filter(product_id=product_id)
    else:
        product_id = ""
    return qs.order_by("-created_at"), {
        "date_from": date_from,
        "date_to": date_to,
        "status": status,
        "product": product_id,
    }


def _ensure_report_permission(request):
    if not (
        request.user.has_perm("_product.view_savingtransaction")
        or request.user.has_perm("_product.change_savingtransaction")
    ):
        raise PermissionDenied


@user_passes_test(_report_permission)
def saving_report(request):
    _ensure_report_permission(request)
    savings, filters = _filtered_savings(request)
    totals = savings.aggregate(count=Count("pk"), amount=Sum("total_amount"))
    paid = savings.filter(status=SavingTransaction.Status.PAID)
    query = request.GET.copy()
    query.pop("page", None)
    context = {
        "rows": Paginator(savings, 50).get_page(request.GET.get("page")),
        "transaction_count": totals["count"] or 0,
        "total_amount": totals["amount"] or 0,
        "member_count": savings.values("email").distinct().count(),
        "paid_count": paid.count(),
        "filters": filters,
        "products": Product.objects.filter(category__slug="simpanan").order_by("title"),
        "statuses": SavingTransaction.Status.choices,
        "export_query": query.urlencode(),
        "transaction_list_url": reverse("wagtailsnippets__product_savingtransaction:list"),
    }
    return render(request, "_product/admin/saving_report.html", context)


def _csv_safe(value):
    text = "" if value is None else str(value)
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def _export_rows(savings):
    for item in savings:
        yield [
            item.transaction_code,
            timezone.localtime(item.created_at).strftime("%Y-%m-%d %H:%M"),
            item.full_name,
            item.nomor_anggota or "Anggota baru",
            item.phone,
            item.email,
            item.product.title,
            float(item.amount),
            float(item.service_fee),
            float(item.total_amount),
            item.get_payment_channel_display(),
            item.get_status_display(),
            timezone.localtime(item.paid_at).strftime("%Y-%m-%d %H:%M") if item.paid_at else "",
        ]


EXPORT_HEADERS = [
    "Kode Setoran", "Tanggal Transaksi", "Nama", "Nomor Anggota", "WhatsApp", "Email",
    "Produk", "Nominal", "Biaya Layanan", "Total", "Kanal", "Status", "Tanggal Dibayar",
]


@user_passes_test(_report_permission)
def export_saving_report(request, file_format):
    _ensure_report_permission(request)
    savings, filters = _filtered_savings(request)
    filename = f"laporan-tabungan-{filters['date_from']}-{filters['date_to']}"
    if file_format == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        response.write("\ufeff")
        writer = csv.writer(response)
        writer.writerow(EXPORT_HEADERS)
        for row in _export_rows(savings):
            writer.writerow([_csv_safe(value) for value in row])
        return response
    if file_format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Laporan Tabungan"
        sheet.freeze_panes = "A2"
        sheet.append(EXPORT_HEADERS)
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="005DAA")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="center")
        for row in _export_rows(savings):
            sheet.append([_csv_safe(value) if isinstance(value, str) else value for value in row])
        sheet.auto_filter.ref = sheet.dimensions
        widths = [22, 19, 25, 18, 18, 28, 24, 16, 16, 16, 22, 18, 19]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        for row in sheet.iter_rows(min_row=2, min_col=8, max_col=10):
            for cell in row:
                cell.number_format = '"Rp" #,##0'
        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        return response
    if file_format == "pdf":
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        output = BytesIO()
        document = SimpleDocTemplate(
            output, pagesize=landscape(A4), rightMargin=10 * mm, leftMargin=10 * mm,
            topMargin=12 * mm, bottomMargin=12 * mm,
        )
        styles = getSampleStyleSheet()
        data = [["Tanggal", "Kode", "Nama / Anggota", "Produk", "Total", "Kanal", "Status"]]
        for item in savings:
            data.append([
                timezone.localtime(item.created_at).strftime("%d-%m-%Y %H:%M"),
                item.transaction_code,
                f"{item.full_name}\n{item.nomor_anggota or 'Anggota baru'}",
                item.product.title,
                f"Rp{item.total_amount:,.0f}".replace(",", "."),
                item.get_payment_channel_display(),
                item.get_status_display(),
            ])
        table = Table(data, repeatRows=1, colWidths=[28*mm, 43*mm, 45*mm, 42*mm, 28*mm, 34*mm, 28*mm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#005DAA")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), .3, colors.HexColor("#DCE2E8")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FB")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story = [
            Paragraph("Laporan Setoran Simpanan", styles["Title"]),
            Paragraph(f"Periode {filters['date_from']:%d-%m-%Y} sampai {filters['date_to']:%d-%m-%Y}", styles["Normal"]),
            Spacer(1, 6*mm), table,
        ]
        document.build(story)
        response = HttpResponse(output.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
        return response
    raise PermissionDenied
